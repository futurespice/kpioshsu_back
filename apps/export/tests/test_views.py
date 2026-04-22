from io import BytesIO

from openpyxl import load_workbook
from rest_framework.test import APITestCase
from rest_framework_simplejwt.tokens import RefreshToken

from apps.departments.models import Department
from apps.faculties.models import Faculty
from apps.users.models import Role, User


class _AuthMixin:
    def _auth(self, user):
        token = RefreshToken.for_user(user).access_token
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")


class _BaseExportTests(_AuthMixin, APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            email="admin@oshsu.kg", password="p", role=Role.ADMIN
        )
        self.teacher = User.objects.create_user(
            email="t@oshsu.kg", password="p", role=Role.TEACHER
        )
        self.dean = User.objects.create_user(
            email="dean@oshsu.kg", password="p", role=Role.DEAN
        )
        self.rector = User.objects.create_user(
            email="r@oshsu.kg", password="p", role=Role.RECTOR
        )
        self.faculty = Faculty.objects.create(name="Инж")
        self.dept = Department.objects.create(name="ИТ", faculty=self.faculty)


class TeacherKPIExportTests(_BaseExportTests):
    @property
    def url(self):
        return f"/api/export/kpi/teacher/{self.teacher.id}/"

    def test_unauthenticated_returns_401(self):
        self.assertEqual(self.client.get(self.url).status_code, 401)

    def test_xlsx_returns_file(self):
        self._auth(self.teacher)
        res = self.client.get(self.url)
        self.assertEqual(res.status_code, 200)
        self.assertIn(
            "openxmlformats-officedocument.spreadsheetml",
            res["Content-Type"],
        )
        self.assertIn("attachment", res["Content-Disposition"])
        # Parse XLSX
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(
            headers, ["Показатель", "Значение", "Вес", "Баллы"]
        )

    def test_pdf_returns_501(self):
        self._auth(self.teacher)
        res = self.client.get(self.url, {"format": "pdf"})
        self.assertEqual(res.status_code, 501)

    def test_invalid_format_returns_422(self):
        self._auth(self.teacher)
        res = self.client.get(self.url, {"format": "docx"})
        self.assertEqual(res.status_code, 422)


class DepartmentKPIExportTests(_BaseExportTests):
    def test_teacher_forbidden(self):
        self._auth(self.teacher)
        res = self.client.get(f"/api/export/kpi/department/{self.dept.id}/")
        self.assertEqual(res.status_code, 403)

    def test_dean_can_export(self):
        self._auth(self.dean)
        res = self.client.get(f"/api/export/kpi/department/{self.dept.id}/")
        self.assertEqual(res.status_code, 200)


class FacultyKPIExportTests(_BaseExportTests):
    def test_rector_can_export(self):
        self._auth(self.rector)
        res = self.client.get(f"/api/export/kpi/faculty/{self.faculty.id}/")
        self.assertEqual(res.status_code, 200)

    def test_teacher_forbidden(self):
        self._auth(self.teacher)
        res = self.client.get(f"/api/export/kpi/faculty/{self.faculty.id}/")
        self.assertEqual(res.status_code, 403)


class TasksExportTests(_BaseExportTests):
    def test_authenticated_can_export(self):
        self._auth(self.teacher)
        res = self.client.get("/api/export/tasks/")
        self.assertEqual(res.status_code, 200)

    def test_pdf_returns_501(self):
        self._auth(self.teacher)
        res = self.client.get("/api/export/tasks/", {"format": "pdf"})
        self.assertEqual(res.status_code, 501)


class PublicationsExportTests(_BaseExportTests):
    def test_authenticated_can_export(self):
        self._auth(self.teacher)
        res = self.client.get(f"/api/export/publications/{self.teacher.id}/")
        self.assertEqual(res.status_code, 200)


class UniversityReportExportTests(_BaseExportTests):
    def test_dean_forbidden(self):
        self._auth(self.dean)
        res = self.client.get("/api/export/report/university/")
        self.assertEqual(res.status_code, 403)

    def test_rector_can_export(self):
        self._auth(self.rector)
        res = self.client.get("/api/export/report/university/")
        self.assertEqual(res.status_code, 200)
